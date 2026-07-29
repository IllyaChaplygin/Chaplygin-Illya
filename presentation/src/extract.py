"""Extract per-SKU self-cost by logistics scenario from SelfCost.xlsx into data.json."""
import json

import openpyxl

SRC = ('/root/.claude/uploads/c53f34b4-ae43-52db-aeb2-01bdcda48cc3/'
       '2dedbaa6-SelfCost.xlsx')
OUT = '/home/user/Chaplygin-Illya/presentation/src/data.json'

NAME, USD, UAH, RATE = 'C', 'AN', 'AO', 'AJ'

# The Retail+Bulk sheets swap the first two columns: there the supplier price sits
# in D and the pack count in E, the other way round from the Retail sheets.
FOB_COL = {True: 'D', False: 'E'}

# sheet -> ({first data row of each block: scenario}, SKUs per block)
LAYOUT = {
    'SINGHA KAMEDA - Retail': (
        {6: "20'", 15: "40'", 24: 'LCL 17', 33: 'LCL 34'}, 2),
    'SINGHA KAMEDA - Retail+Bulk': (
        {6: "20'", 17: "40'", 28: 'LCL 17', 39: 'LCL 34'}, 4),
    'Thai-Nichi - Retail': (
        {6: "20'", 15: "40'", 24: 'LCL 17', 32: 'LCL 34'}, 2),
    'Thai-Nichi - Retail+Bulk': (
        {6: "20'", 17: "40'", 28: 'LCL 17', 39: 'LCL 34'}, 4),
    'TMK Thailand Co., Ltd -Retail': (
        {6: "20'", 22: "40'", 39: 'LCL 17', 56: 'LCL 34'}, 10),
    'ZEK -Retail': (
        {6: "20'", 26: "40'", 46: 'LCL 17', 66: 'LCL 34'}, 14),
}


def clean(s):
    return ' '.join(str(s).replace('•', '').replace('\t', ' ').split())


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    out = {}
    for sheet, (blocks, span) in LAYOUT.items():
        ws = wb[sheet]
        fob = FOB_COL[sheet.endswith('Bulk')]
        skus, order = {}, []
        for start, scenario in blocks.items():
            for i in range(span):
                r = start + i
                name = ws[NAME + str(r)].value
                assert name and 'аименование' not in str(name), (sheet, r, name)
                name = clean(name)
                if name not in skus:
                    skus[name] = {'name': name, 'fob': ws[fob + str(r)].value,
                                  'cost': {}}
                    order.append(name)
                skus[name]['cost'][scenario] = {
                    'usd': round(ws[USD + str(r)].value, 4),
                    'uah': round(ws[UAH + str(r)].value, 2),
                    'rate': ws[RATE + str(r)].value,
                }
        assert len(order) == span, (sheet, order)
        out[sheet] = [skus[n] for n in order]

    with open(OUT, 'w') as f:
        json.dump(out, f, ensure_ascii=False, indent=1)

    for sheet, rows in out.items():
        print('=' * 78)
        print(sheet, len(rows), 'SKU')
        for r in rows:
            print('  %-42s | ' % r['name'][:42] +
                  ' | '.join('%s $%.3f / %.2f₴' % (k, v['usd'], v['uah'])
                             for k, v in r['cost'].items()))


if __name__ == '__main__':
    main()
