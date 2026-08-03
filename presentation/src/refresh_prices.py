"""Refresh data.json prices from a new SelfCost revision, keeping SKU names.

The 2026-08 SelfCost revision labels the same product differently between
container blocks (e.g. Singha "Norimaki Original 42 g" in the 20'/40' blocks
but "Original Flavour" in the LCL blocks; TMK rolls carry a size suffix in
some blocks and not others). Matching by name therefore breaks, so prices are
pulled positionally — product i of a block is the same product across all four
blocks — while the display name is kept from the existing data.json so the
catalog `key`s still resolve.

Usage:  python3 src/refresh_prices.py  [path/to/SelfCost.xlsx]
"""
import json
import os
import sys

import openpyxl

HERE = os.path.dirname(__file__)
DATA = os.path.join(HERE, 'data.json')
DEFAULT_SRC = ('/root/.claude/uploads/c53f34b4-ae43-52db-aeb2-01bdcda48cc3/'
               '50d60843-SelfCost.xlsx')

USD, UAH, RATE = 'AN', 'AO', 'AJ'
FOB_COL = {True: 'D', False: 'E'}      # Retail+Bulk keeps the FOB price in D

# sheet -> ({first data row of each block: scenario}, SKUs per block)
LAYOUT = {
    'SINGHA KAMEDA - Retail': ({6: "20'", 15: "40'", 24: 'LCL 17', 33: 'LCL 34'}, 2),
    'SINGHA KAMEDA - Retail+Bulk': ({6: "20'", 17: "40'", 28: 'LCL 17', 39: 'LCL 34'}, 4),
    'Thai-Nichi - Retail': ({6: "20'", 15: "40'", 24: 'LCL 17', 32: 'LCL 34'}, 2),
    'Thai-Nichi - Retail+Bulk': ({6: "20'", 17: "40'", 28: 'LCL 17', 39: 'LCL 34'}, 4),
    'TMK Thailand Co., Ltd -Retail': ({6: "20'", 22: "40'", 39: 'LCL 17', 56: 'LCL 34'}, 10),
    'ZEK -Retail': ({6: "20'", 26: "40'", 46: 'LCL 17', 66: 'LCL 34'}, 14),
}


def main(src=DEFAULT_SRC):
    data = json.load(open(DATA, encoding='utf-8'))
    wb = openpyxl.load_workbook(src, data_only=True)
    changed = 0
    for sheet, (blocks, span) in LAYOUT.items():
        ws = wb[sheet]
        fob = FOB_COL[sheet.endswith('Bulk')]
        assert len(data[sheet]) == span, (sheet, len(data[sheet]), span)
        first = next(iter(blocks))
        for i in range(span):
            item = data[sheet][i]
            item['fob'] = ws[fob + str(first + i)].value
            cost = {}
            for start, scen in blocks.items():
                r = start + i
                cost[scen] = {'usd': round(ws[USD + str(r)].value, 4),
                              'uah': round(ws[UAH + str(r)].value, 2),
                              'rate': ws[RATE + str(r)].value}
            if cost != item['cost']:
                changed += 1
            item['cost'] = cost
    json.dump(data, open(DATA, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
    print('refreshed %s -> data.json (%d SKUs with changed prices)'
          % (os.path.basename(src), changed))


if __name__ == '__main__':
    main(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC)
